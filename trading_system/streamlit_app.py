"""
Streamlit Dashboard for CFD Trading Signal System
Optimized version: Fast Mode only, no unnecessary code
"""

import streamlit as st
import pandas as pd
import os
import sys
from pathlib import Path
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Add project to path
PROJECT_ROOT = Path(__file__).parent
sys.path.insert(0, str(PROJECT_ROOT))

try:
    from analysis.indicators import IndicatorCalculator
    from analysis.sentiment import SentimentAnalyzer
    from analysis.earnings import EarningsChecker
    from agents.crew import CFDTradingCrew
    from utils.trade_card import TradeCardBuilder, TradeCardWriter
except ImportError:
    from trading_system.analysis.indicators import IndicatorCalculator
    from trading_system.analysis.sentiment import SentimentAnalyzer
    from trading_system.analysis.earnings import EarningsChecker
    from trading_system.agents.crew import CFDTradingCrew
    from trading_system.utils.trade_card import TradeCardBuilder, TradeCardWriter

import yaml
from dotenv import load_dotenv

# Load environment
load_dotenv(PROJECT_ROOT / '.env')

# Page config
st.set_page_config(
    page_title="CFD Trading Signal System",
    page_icon="📈",
    layout="wide"
)

st.markdown("""
<style>
    .buy-signal { background-color: #d4edda; padding: 15px; border-left: 4px solid #28a745; border-radius: 5px; margin: 10px 0; }
    .sell-signal { background-color: #f8d7da; padding: 15px; border-left: 4px solid #dc3545; border-radius: 5px; margin: 10px 0; }
    .hold-signal { background-color: #fff3cd; padding: 15px; border-left: 4px solid #ffc107; border-radius: 5px; margin: 10px 0; }
</style>
""", unsafe_allow_html=True)

# Initialize session state
if 'system_initialized' not in st.session_state:
    st.session_state.system_initialized = False
    st.session_state.last_results = None
    st.session_state.running = False

# Load configuration
@st.cache_resource
def load_config():
    config_path = PROJECT_ROOT / 'config.yaml'
    with open(config_path, 'r') as f:
        return yaml.safe_load(f)

@st.cache_resource
def init_system():
    config = load_config()
    return {
        'config': config,
        'indicator_calc': IndicatorCalculator(config),
        'sentiment_analyzer': SentimentAnalyzer(config),
        'earnings_checker': EarningsChecker(config),
        'trade_card_builder': TradeCardBuilder(config),
        'trade_card_writer': TradeCardWriter(config),
        'crew': CFDTradingCrew(config)
    }

# Helper: Export to Excel
def export_to_excel(trade_cards_data):
    """Convert trade cards to Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Trade Signals"
    
    headers = ["Ticker", "Signal", "Confidence", "Entry", "Stop Loss", "TP1", "TP2", "Leverage", "Sentiment"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for card in trade_cards_data:
        signal = card.get('signal', '')
        row = [
            card.get('ticker', ''),
            signal,
            card.get('confidence', 0),
            card.get('entry_price', 0),
            card.get('stop_loss', 0),
            card.get('take_profit_1', 0),
            card.get('take_profit_2', 0),
            f"{card.get('leverage_recommended', 0)}:1",
            card.get('sentiment_score', 0),
        ]
        ws.append(row)
        
        if signal == 'BUY':
            fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif signal == 'SELL':
            fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        else:
            fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        for cell in ws[ws.max_row]:
            cell.fill = fill
    
    # Adjust columns
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 12
    
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()

# ==================== MAIN UI ====================

st.header("🚀 CFD Trading Signal System")

tab1, tab2, tab3 = st.tabs(["🚀 Daily Scan", "📊 View Results", "❓ About"])

# TAB 1: SCAN
with tab1:
    st.subheader("⚡ Quick Scan (2-5 minutes)")
    st.markdown("""
    **How it works:**
    1. Downloads price data for all tickers (parallel)
    2. Calculates technical indicators (RSI, MACD, ATR)
    3. Gets sentiment from news
    4. **If signal is BUY/SELL**: Runs CrewAI to confirm
    5. **If signal is HOLD**: Skips AI (saves time)
    """)
    
    if st.button("▶️ Start Scan", key="daily_scan", help="Quick technical scan"):
        st.session_state.running = True
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        system = init_system()
        config = system['config']
        watchlist = config.get('watchlist', [])
        
        all_trade_cards = []
        
        # STEP 1: Parallel download
        status_text.text("📥 Downloading data...")
        daily_data = system['indicator_calc'].get_daily_data_parallel(watchlist, max_workers=5)
        
        all_indicators = {}
        for ticker in watchlist:
            try:
                data = daily_data.get(ticker)
                if data is not None:
                    indicators_data = system['indicator_calc'].calculate_all_indicators_from_data(ticker, data)
                    if indicators_data:
                        all_indicators[ticker] = indicators_data
            except Exception as e:
                print(f"Error: {ticker}: {e}")
        
        # STEP 2: Process each ticker
        processed = 0
        for ticker in watchlist:
            if ticker not in all_indicators:
                continue
            
            processed += 1
            progress = processed / len(watchlist)
            progress_bar.progress(progress)
            status_text.text(f"Analyzing {ticker}... ({processed}/{len(watchlist)})")
            
            try:
                indicators_data = all_indicators[ticker]
                
                # Check earnings
                skip_earnings, earnings_date = system['earnings_checker'].check_earnings_within_threshold(ticker)
                if skip_earnings:
                    trade_card = system['trade_card_builder'].build_trade_card(
                        ticker=ticker,
                        signal="HOLD",
                        current_price=indicators_data['current_price'],
                        atr=indicators_data['atr'],
                        indicators_data=indicators_data,
                        confidence=0,
                        catalyst=f"Earnings on {earnings_date}",
                        sentiment_score=0.0,
                        skip_reason="EARNINGS"
                    )
                    all_trade_cards.append(trade_card)
                    continue
                
                # Get indicators
                rsi = indicators_data.get('rsi', 50)
                macd_cross = indicators_data.get('macd_cross', '')
                
                # FAST signal generation
                if rsi < 35 and macd_cross == 'bullish':
                    fast_signal = "BUY"
                    confidence = 0.7
                elif rsi > 65 and macd_cross == 'bearish':
                    fast_signal = "SELL"
                    confidence = 0.7
                else:
                    fast_signal = "HOLD"
                    confidence = 0.5
                
                # Get sentiment (always cached)
                sentiment_score = system['sentiment_analyzer'].calculate_sentiment_score(ticker)
                top_headlines = system['sentiment_analyzer'].get_top_headlines(ticker, limit=3)
                
                # ONLY RUN CREW AI IF BUY OR SELL
                if fast_signal in ["BUY", "SELL"]:
                    status_text.text(f"🧠 Confirming {ticker} with AI...")
                    
                    trade_card = system['crew'].run_signal_generation(
                        ticker=ticker,
                        indicators_data=indicators_data,
                        sentiment_score=sentiment_score,
                        top_headlines=top_headlines
                    )
                    
                    if trade_card:
                        all_trade_cards.append(trade_card)
                    else:
                        # Fallback if crew fails
                        trade_card = system['trade_card_builder'].build_trade_card(
                            ticker=ticker,
                            signal=fast_signal,
                            current_price=indicators_data['current_price'],
                            atr=indicators_data['atr'],
                            indicators_data=indicators_data,
                            confidence=confidence,
                            catalyst=f"RSI={rsi} MACD={macd_cross}",
                            sentiment_score=sentiment_score,
                            skip_reason="CREW_ERROR"
                        )
                        all_trade_cards.append(trade_card)
                else:
                    # HOLD: Skip CrewAI entirely
                    status_text.text(f"⏭️ {ticker} = HOLD (skipping AI)")
                    trade_card = system['trade_card_builder'].build_trade_card(
                        ticker=ticker,
                        signal="HOLD",
                        current_price=indicators_data['current_price'],
                        atr=indicators_data['atr'],
                        indicators_data=indicators_data,
                        confidence=confidence,
                        catalyst=f"RSI={rsi} MACD={macd_cross}",
                        sentiment_score=sentiment_score,
                        skip_reason="NO_SETUP"
                    )
                    all_trade_cards.append(trade_card)
            
            except Exception as e:
                st.warning(f"Error {ticker}: {e}")
        
        progress_bar.progress(1.0)
        status_text.text("✅ Scan complete!")
        
        if all_trade_cards:
            system['trade_card_writer'].write_trade_cards(all_trade_cards)
            st.session_state.last_results = all_trade_cards
        
        st.session_state.running = False

# TAB 2: RESULTS
with tab2:
    st.subheader("📊 Latest Results")
    
    if st.session_state.last_results:
        results_df = pd.DataFrame([
            {
                'Ticker': card.get('ticker', ''),
                'Signal': card.get('signal', ''),
                'Confidence': f"{card.get('confidence', 0):.0%}",
                'Entry': f"${card.get('entry_price', 0):.2f}",
                'Stop Loss': f"${card.get('stop_loss', 0):.2f}",
                'RSI': f"{card.get('rsi', 0):.0f}",
                'MACD': card.get('macd_cross', ''),
                'Sentiment': f"{card.get('sentiment_score', 0):.2f}",
            }
            for card in st.session_state.last_results
        ])
        
        st.dataframe(results_df, use_container_width=True)
        
        # Export button
        if st.button("📥 Export to Excel"):
            excel_file = export_to_excel(st.session_state.last_results)
            st.download_button(
                label="Download Excel",
                data=excel_file,
                file_name=f"signals_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.info("No results yet. Run a scan first!")

# TAB 3: ABOUT
with tab3:
    st.markdown("""
    ## About
    
    **CFD Trading Signal System**
    
    Uses technical indicators + AI to generate trading signals:
    - **RSI** (oversold/overbought)
    - **MACD** (momentum confirmation)
    - **Sentiment** (news analysis)
    - **CrewAI** (confirmation on strong signals only)
    
    **Fast Mode**: Only runs AI on BUY/SELL signals → 2-5 minutes
    
    ---
    Made with ❤️ using Streamlit + CrewAI
    """)
